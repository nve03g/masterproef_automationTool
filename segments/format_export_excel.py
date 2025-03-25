import pandas as pd
import warnings
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill # export font data to Excel
from openpyxl.styles.colors import Color, RGB # for converting rgb color value to a Color object

import xlwings as xw

warnings.simplefilter("ignore", UserWarning) # we krijgen warning dat openpyxl geen dropdownlijsten in excel meer ondersteunt, maar dat is geen probleem want die controle ga ik via mijn python code uitvoeren, dus deze warning mag genegeerd worden

class ExcelDataProcessor:
    def __init__(self, filepath):

        self.filepath = filepath
        self.data_dfs = {} # dictionary containing all data {sheetname: DataFrame}
        self.format_dfs = {} # dictionary containing all format data {sheetname: DataFrame}
        
    def load_all_sheets(self):
        """ Laad de Excel-sheets in dataframes. """
        try:
            dfs = pd.read_excel(self.filepath, sheet_name=None, header=None) # all sheets, without headers
            self.data_dfs = {sheet: df for sheet, df in dfs.items()}
        except Exception as e:
            print(f"Fout bij laden van de sheets: {e}")
                
    def load_formatting(self):
        """ Laad formattering per cel in dicts in dfs. """
        try:
            wb_xl = load_workbook(self.filepath)
            wb_xw = xw.Book(self.filepath)
            for sheet in wb_xl.sheetnames:
                ws_xl = wb_xl[sheet]
                ws_xw = wb_xw[sheet]
                
                format_data = []
                
                for row in ws_xl.iter_rows(min_row=1, max_row=ws_xl.max_row, min_col=1, max_col=ws_xl.max_column):
                    format_row = [
                        {
                            "cell color": ws_xw.range((cell.row, cell.column)).color, # geeft een tuple (R, G, B) of None als geen kleur ingesteld
                            "text color": ws_xw.range((cell.row, cell.column)).font.color, # geeft een tuple (R, G, B), (0, 0, 0) als geen kleur ingesteld (default zwarte tekst)
                            "font": cell.font.name,
                            "bold": cell.font.bold,
                            "italic": cell.font.italic,
                            "strikethrough": cell.font.strike
                        }
                        for cell in row
                    ]
                    format_data.append(format_row)

                self.format_dfs[sheet] = pd.DataFrame(format_data)

        except Exception as e:
            print(f"Fout bij laden van de formattering: {e}")
        
    def get_dataframe(self, sheetname):
        """ get specific sheet (DataFrame) """
        return self.data_dfs.get(sheetname)
    
    def export_new_excel(self, output_filepath):
        """ Save edited data and formatting to a new Excel file. """
        try:
            wb = Workbook() # make new Excel workbook
            
            for sheetname, df in self.data_dfs.items():
                ws = wb.create_sheet(title=sheetname) # make new sheet
                
                # add data to the sheet
                for row_index, row in df.iterrows():
                    for column_index, value in enumerate(row):
                        cell = ws.cell(row=row_index+1, column=column_index+1, value=value)
                        
                        # pas formattering toe
                        if sheetname in self.format_dfs:
                            format_data = self.format_dfs[sheetname].iloc[row_index, column_index]
                            print(type(format_data["text color"]))
                            print(format_data["text color"])
                            
                            font = Font(
                                # bold=format_data["bold"],
                                # italic=format_data["italic"],
                                # strike=format_data["strikethrough"],
                                # color=format_data["text color"] # tekstkleur
                                color="00FF8000"
                            )
                            cell.font = font

                            # Kleurinstellingen aanpassen:
                            if format_data["cell color"]:
                                # Zorgen dat de cell color een hex string is, bijvoorbeeld 'FF0000'
                                cell.fill = PatternFill(start_color=format_data["cell color"], end_color=format_data["cell color"], fill_type="solid")
            
            # verwijder by default aangemaakte "Sheet"
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
                
            # save workbook to file
            wb.save(output_filepath)
            print(f"New Excel-file saved as {output_filepath}")
            
        except Exception as e:
            print(f"Fout bij opslaan naar nieuwe Excel: {e}")
    

file_path = "format_test.xlsx"
output_file_path = "format_export_excel_outputted.xlsx"
processor = ExcelDataProcessor(file_path)
processor.load_all_sheets()
processor.load_formatting()

# processor.export_new_excel(output_file_path)

data_dictionary = processor.data_dfs
format_dictionary = processor.format_dfs

print(data_dictionary)
print()
print(format_dictionary)

# print format data per cell in certain sheet
# sheetname = "Sheet2"
# first_row_format_df = format_dictionary.get("Sheet1").loc[0] # type: dict
# for i in range(data_dictionary.get(sheetname).shape[0]):  # Iterate through rows
#     for j in range(data_dictionary.get(sheetname).shape[1]): # iterate through columns
# for key in data_dictionary:
#     print(f"{key}")
# print("\n")

