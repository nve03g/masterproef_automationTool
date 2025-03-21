import xlwings as xw 
# columns are accessed by their numeric index, starting from 1 (dus niet ABC maar 123)
# => ws["B4"] = ws.cells(4, 2) (row, column)
import openpyxl as xl

def max_char(filename, sheetname, column_name, max_chars):
    wb = xw.Book(filename) # open Excel file
    ws = wb.sheets[sheetname] # select the right sheet
    
    # read headers (hier in rij 3)
    ## Nellie: dit nog parametrizeren, headers niet in alle files op deze plaats
    headers = {cel.value: cel.column for cel in ws.range("3:3")}  # Dict {column name: column letter}

    # check if column exists
    if column_name not in headers:
        print(f"Column '{column_name}' not found.")
        wb.close()
        return

    column_letter = headers[column_name]  # fetch column letter ('A', 'B' ...)
    errors = []
    
    # identify the last row
    last_row = ws.used_range.last_cell.row

    # run through cells in column (start at row 4, first row of data after headers)
    # Nellie: dit nog parametrizeren, afhankelijk van waar headers staan in file (niet in elke file hetzelfde)
    for row in range(4, last_row + 1):  
        cell_value = ws.cells[row, column_letter].value
        if isinstance(cell_value, str) and len(cell_value) > max_chars:
            errors.append(f"Row {row}: '{cell_value}' is too long ({len(cell_value)} > {max_chars})")

    # show results
    # Nellie: errors niet per se printen, maar opslaan in een logfile of dergelijke
    if errors:
        print("Errors found:")
        for error in errors:
            print(error)
    else:
        print(f"No errors found in column '{column_name}'")
        
    wb.close() # no changes made to file, no need to save


def alarm_exists(ws, alarm_column_index, row_nr):
    cell_value = ws.cells[row_nr, alarm_column_index] # access specific cell to check if alarm exists in this row
    
    # non existing -> empty, "reserved" or "spare"
    if cell_value in [None, "", "reserved", "spare"]:
        return False
    return True
            

def empty(filename, sheetname, column_name, alarm_column_name, must_be_empty=True):
    """
    verbetering:
    kunnen we aan de functie empty() een condition meegeven? dan kan alarm_exists() een mogelijke conditie zijn, en ook andere condities (zie checklist)
    """
    wb = xl.load_workbook(filename) # load Excel file
    ws = wb[sheetname] # access the right sheet
    
    # read headers
    ## Nellie: dit nog parametrizeren, headers niet in alle files op deze plaats
    headers = {cel.value: cel.column_letter for cel in ws[3]}  # Dict {column name: column letter}

    # check if column exists
    if column_name not in headers:
        print(f"Column '{column_name}' not found.")
        return

    column_letter = headers[column_name]  # fetch column letter ('A', 'B' ...)
    errors = []

    # run through cells in column (start at row 4, first row of data after headers)
    # Nellie: dit nog parametrizeren, afhankelijk van waar headers staan in file (niet in elke file hetzelfde)
    for row, cell in enumerate(ws[column_letter][3:], start=4): # verloopt traag omdat je over elke rij loopt (800+ rijen)
        if alarm_exists(filename, sheetname, alarm_column_name, row):
            cell_content = cell.value
            
            # check if cell must be empty or not (based on parameter)
            if must_be_empty and cell_content not in [None, ""]:
                errors.append(f"Row {row}: '{cell_content}' must be empty, but contains a value!")
            elif not must_be_empty and cell_content in [None, ""]:
                errors.append(f"Row {row}: This field cannot be empty!")
            else:
                # print(f"Row {row}: The cell complies with the empty-check.")
                pass
        else:
            # print(f"Row {row}: Alarm does not exist, skipping empty-check.")
            pass
        
    # show results
    # Nellie: errors niet per se printen, maar opslaan in een logfile of dergelijke
    if errors:
        print("Errors found:")
        for error in errors:
            print(error)
    else:
        print(f"No errors found in column '{column_name}'")


def empty_with_correction(filename, sheetname, column_name, alarm_column_name, must_be_empty=True):
    """
    verbetering:
    kunnen we aan de functie empty_with_correction() een condition meegeven? dan kan alarm_exists() een mogelijke conditie zijn, en ook andere condities (zie checklist)
    """
    wb = xl.load_workbook(filename) # load Excel file
    ws = wb[sheetname] # access the right sheet
    
    # read headers
    ## Nellie: dit nog parametrizeren, headers niet in alle files op deze plaats
    headers = {cel.value: cel.column_letter for cel in ws[3]}  # Dict {column name: column letter}

    # check if column exists
    if column_name not in headers:
        print(f"Column '{column_name}' not found.")
        return

    column_letter = headers[column_name]  # fetch column letter ('A', 'B' ...)
    correction_count = 0

    # run through cells in column (start at row 4, first row of data after headers)
    # Nellie: dit nog parametrizeren, afhankelijk van waar headers staan in file (niet in elke file hetzelfde)
    for row in range(4, ws.max_row + 1): # verloopt traag omdat je over elke rij loopt (800+ rijen)
        if alarm_exists(filename, sheetname, alarm_column_name, row): # data control
            cell = ws[f"{column_letter}{row}"]
            cell_content = cell.value

            # data correction
            if must_be_empty and cell_content not in [None, ""]:
                cell.value = ""  # correct to an empty cell
                correction_count += 1
            elif not must_be_empty and cell_content in [None, ""]:
                cell.value = "emptiness correction"  # here we have to ask to fill in the field
                correction_count += 1

    # save corrections as new file
    if correction_count > 0:
        wb.save(f"{filename}_corrected.xlsx")
        print(f"Corrected file saved as '{f"{filename}_corrected.xlsx"}' with {correction_count} changes.")
    else:
        print("No corrections were necessary.")


def file_type(filename, sheetname, column_name, alarm_column_name, file_extension=".pdl"):
    wb = xw.Book(filename) # load Excel file
    ws = wb.sheets[sheetname] # access the right sheet
    
    # read headers
    ## Nellie: dit nog parametrizeren, headers niet in alle files op deze plaats
    headers = {cel.value: cel.column for cel in ws.range("3:3")}  # Dict {column name: column letter}

    # check if column exists
    if column_name not in headers:
        print(f"Column '{column_name}' not found.")
        wb.close()
        return
    
    if alarm_column_name not in headers:
        print(f"Column '{alarm_column_name}' not found.")
        wb.close()
        return

    column_index = headers[column_name]  # fetch column letter ('A', 'B' ...)
    alarm_column_index = headers[alarm_column_name]
    errors = []
    
    # identify last row
    last_row = ws.used_range.last_cell.row

    # run through cells in column (start at row 4, first row of data after headers)
    # Nellie: dit nog parametrizeren, afhankelijk van waar headers staan in file (niet in elke file hetzelfde)
    for row in range(4, last_row + 1):
        if alarm_exists(ws, alarm_column_index, row): # Nellie: is dit een vereiste?
            cell_value = str(ws.cells(row, column_index).value).strip() if ws.cells(row, column_index).value else ""
            
            # print(f"Row {row}, Column {column_name} (Index {column_index}): {cell_value}")

            # check if the value ends with the required file extension
            if not cell_value.endswith(file_extension):
                errors.append(f"Row {row}: '{cell_value}' is not a valid {file_extension} file.")
        else:
            pass

    # show results
    # Nellie: errors niet per se printen, maar opslaan in een logfile of dergelijke
    if errors:
        print("Errors found:")
        for error in errors:
            print(error)
    else:
        print(f"No errors found in column '{column_name}'")
        
    # wb.close()


def input_select(filename, sheetname, column_name, alarm_column_name, mode, entries):
    """
    verbetering:
    lijst van entries via constante of config file vastleggen, zodat we die niet allemaal rechtstreeks moeten invullen in function call
    """
    wb = xl.load_workbook(filename) # load Excel file
    ws = wb[sheetname] # access the right sheet
    
    # read headers
    ## Nellie: dit nog parametrizeren, headers niet in alle files op deze plaats
    headers = {cel.value: cel.column_letter for cel in ws[3]}  # Dict {column name: column letter}

    # check if column exists
    if column_name not in headers:
        print(f"Column '{column_name}' not found.")
        return

    column_letter = headers[column_name]  # fetch column letter ('A', 'B' ...)
    errors = []

    # run through cells in column (start at row 4, first row of data after headers)
    # Nellie: dit nog parametrizeren, afhankelijk van waar headers staan in file (niet in elke file hetzelfde)
    for row in range(4, ws.max_row + 1):
        if alarm_exists(filename, sheetname, alarm_column_name, row):
            cell = ws[f"{column_letter}{row}"]
            cell_value = str(cell.value).strip() if cell.value else ""

            # check input based on mode (parameter)
            if mode == 1: # can contain only one of the entries
                if cell_value not in entries:
                    errors.append(f"Row {row}: '{cell_value}' is not a valid entry. Must be one of {entries}.")
            elif mode == 0: # can contain one or more of the entries
                if not any(entry in cell_value for entry in entries):
                    errors.append(f"Row {row}: '{cell_value}' must contain at least one of {entries}.")
            else:
                print("Invalid mode. Use 1 for exact match, 0 for partial match.")
                return
        else:
            pass

    # show results
    # Nellie: errors niet per se printen, maar opslaan in een logfile of dergelijke
    if errors:
        print(f"Errors found in column '{column_name}':")
        for error in errors:
            print(error)
    else:
        print(f"All values in column '{column_name}' comply with the check.")


### test the functions
# max_char(filename="voorbeeld_alarmlijst.xlsx", sheetname="Alarmlist", column_name="Alarmtext English", max_chars=75)

empty("voorbeeld_alarmlijst.xlsx", "Alarmlist", "Class", "Alarmtext machine constructor (German)", must_be_empty=False)
# empty("voorbeeld_alarmlijst.xlsx", "Alarmlist", "Pass / fail", "Alarmtext machine constructor (German)", must_be_empty=True)

# empty_with_correction("voorbeeld_alarmlijst.xlsx", "Alarmlist", "Class", "Alarmtext machine constructor (German)", must_be_empty=False)

# file_type(filename="voorbeeld_alarmlijst.xlsx", sheetname="Alarmlist", column_name="Picture", alarm_column_name="Alarmtext machine constructor (German)", file_extension=".pdl")

# nog te testen!
# input_select("voorbeeld_alarmlijst.xlsx", "Alarmlist", "Class", "Alarmtext machine constructor (German)", 1, ["A3", "A4", "A5", "B", "C1", "C2", "C3", "D1", "D2", "D3"])